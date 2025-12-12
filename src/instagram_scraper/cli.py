from __future__ import annotations

import argparse
import json
import logging
from pathlib import Path

from .config import load_config
from .scraper import InstagramScraper
from .auth import FacebookAuthenticator
from .browser_scraper import BrowserInstagramScraper


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Instagram Scraper con login Facebook (OAuth)")
    subparsers = parser.add_subparsers(dest="command", required=True)

    # Subcomando de autenticación
    auth_parser = subparsers.add_parser("auth", help="Autenticación con Facebook y guardado de sesión")
    auth_parser.add_argument("--headless", type=str, default=None, help="true/false para ejecutar en modo headless")

    # Subcomando de scraping con sesión Playwright
    scrape_parser = subparsers.add_parser("scrape", help="Scrapear perfil usando sesión autenticada")
    scrape_parser.add_argument("--url", required=True, help="Enlace del perfil de Instagram")
    scrape_parser.add_argument("--posts", type=int, default=None, help="Cantidad de posts recientes (por defecto POSTS_LIMIT)")
    scrape_parser.add_argument("--output", type=Path, default=None, help="Archivo de salida JSON (opcional)")

    followers_parser = subparsers.add_parser("followers", help="Listar seguidores y conteos de sus seguidores")
    followers_parser.add_argument("--url", required=True, help="Enlace del perfil de Instagram")
    followers_parser.add_argument("--limit", type=int, default=20, help="Cantidad de seguidores a consultar")
    followers_parser.add_argument("--output", type=Path, default=None, help="Archivo de salida JSON (opcional)")
    followers_parser.add_argument("--page-size", type=int, default=12, help="Tamaño de página para paginación")
    followers_parser.add_argument("--chunk", type=int, default=2, help="Tamaño de lote para consultas de detalle")
    followers_parser.add_argument("--delay-ms", type=int, default=3000, help="Retraso entre páginas/lotes en ms")
    followers_parser.add_argument("--retry-tries", type=int, default=10, help="Intentos de reintento ante 429/0")
    followers_parser.add_argument("--retry-base-ms", type=int, default=2500, help="Base de backoff en ms")

    # Subcomando de seguidos (following) y detalles
    following_parser = subparsers.add_parser("following", help="Listar seguidos del perfil y detalles por usuario")
    following_parser.add_argument("--url", required=True, help="Enlace del perfil de Instagram")
    following_parser.add_argument("--limit", type=int, default=20, help="Cantidad de seguidos a consultar")
    following_parser.add_argument("--output", type=Path, default=None, help="Archivo de salida (.xlsx/.csv recomendado)")
    following_parser.add_argument("--page-size", type=int, default=12, help="Tamaño de página para paginación")
    following_parser.add_argument("--chunk", type=int, default=2, help="Tamaño de lote para consultas de detalle")
    following_parser.add_argument("--delay-ms", type=int, default=3000, help="Retraso entre páginas/lotes en ms")
    following_parser.add_argument("--retry-tries", type=int, default=10, help="Intentos de reintento ante 429/0")
    following_parser.add_argument("--retry-base-ms", type=int, default=2500, help="Base de backoff en ms")
    following_parser.add_argument("--force-ui", action="store_true", help="Forzar modo UI (diálogo de seguidos y scroll)")

    # Subcomando de scraping con Instaloader (opcional)
    legacy_parser = subparsers.add_parser("legacy", help="Scrapear con Instaloader (login IG opcional)")
    legacy_parser.add_argument("--url", required=True, help="Enlace del perfil de Instagram")
    legacy_parser.add_argument("--posts", type=int, default=None, help="Cantidad de posts recientes (por defecto POSTS_LIMIT)")
    legacy_parser.add_argument("--output", type=Path, default=None, help="Archivo de salida JSON (opcional)")
    legacy_parser.add_argument("--login", action="store_true", help="Intentar login IG con usuario/contraseña")

    return parser


def main() -> None:
    config = load_config()
    logging.basicConfig(level=getattr(logging, config.log_level.upper(), logging.INFO), format="%(asctime)s %(levelname)s %(name)s: %(message)s")

    parser = build_parser()
    args = parser.parse_args()

    if args.command == "auth":
        if args.headless is not None:
            config.headless = args.headless.lower() == "true"
        auth = FacebookAuthenticator(config)
        auth.login_with_facebook()
        print("Autenticación completada y sesión guardada.")
        return

    elif args.command == "scrape":
        scraper = BrowserInstagramScraper(config)
        data = scraper.get_profile_data(args.url, posts_limit=args.posts)
    elif args.command == "following":
        import time as _t
        t0 = _t.time()
        scraper = BrowserInstagramScraper(config)
        data = scraper.get_following_details(
            args.url,
            following_limit=args.limit,
            page_size=args.page_size,
            chunk=args.chunk,
            delay_ms=args.delay_ms,
            retry_tries=args.retry_tries,
            retry_base_ms=args.retry_base_ms,
            force_ui=getattr(args, "force_ui", False),
        )
        if data is None:
            data = {"username": None, "following_details": []}
        out_path = getattr(args, "output", None)
        items = data.get("following_details", [])
        if out_path and out_path.suffix.lower() in {".xlsx", ".csv"}:
            rows = []
            print(f"Items scrapeados: {len(items)}")
            for it in items:
                try:
                    uname = it.get("username") or ""
                    fn = it.get("full_name") or ""
                    bio = it.get("biography") or ""
                    followers = it.get("followers")
                    following = it.get("following")
                    print(f"[following] {uname} | nombre='{fn}' | bio_len={len(bio)} | seguidores={followers} | seguidos={following}")
                except Exception:
                    pass
                rows.append([
                    it.get("full_name") or "",
                    it.get("username") or "",
                    it.get("biography") or "",
                    it.get("account_type") or "",
                    it.get("category") or "",
                    it.get("followers") if it.get("followers") is not None else "",
                    it.get("following") if it.get("following") is not None else "",
                    it.get("url") or "",
                ])

            if out_path.suffix.lower() == ".xlsx":
                try:
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "following"
                    ws.append(["nombre", "usuario", "biografia", "tipo_de_cuenta", "categoria", "seguidores", "seguidos", "enlace"])
                    for r in rows:
                        ws.append(r)
                    out_path.parent.mkdir(parents=True, exist_ok=True)
                    wb.save(str(out_path))
                    print(f"Archivo Excel guardado en {out_path}")
                    print(f"Tiempo total: {round(_t.time()-t0, 2)}s")
                    return
                except Exception as e:
                    # Intento adicional: si el archivo está abierto (Windows bloquea con Excel), probar con un nombre alternativo
                    msg = str(e)
                    try:
                        if "Permission denied" in msg or getattr(e, "errno", None) == 13:
                            alt_path = out_path.with_name(out_path.stem + "_v2" + out_path.suffix)
                            from openpyxl import Workbook
                            wb2 = Workbook()
                            ws2 = wb2.active
                            ws2.title = "following"
                            ws2.append(["nombre", "usuario", "biografia", "tipo_de_cuenta", "categoria", "seguidores", "seguidos", "enlace"])
                            for r in rows:
                                ws2.append(r)
                            alt_path.parent.mkdir(parents=True, exist_ok=True)
                            wb2.save(str(alt_path))
                            print(f"Archivo Excel bloqueado, se guardó en {alt_path}")
                            print(f"Tiempo total: {round(_t.time()-t0, 2)}s")
                            return
                    except Exception as e2:
                        print(f"No se pudo escribir Excel alternativo: {e2}.")
                    print(f"No se pudo escribir Excel (.xlsx): {e}. Se imprimirá JSON.")

            if out_path.suffix.lower() == ".csv":
                try:
                    import csv
                    out_path.parent.mkdir(parents=True, exist_ok=True)
                    with out_path.open("w", newline="", encoding="utf-8") as f:
                        writer = csv.writer(f)
                        writer.writerow(["nombre", "usuario", "biografia", "tipo_de_cuenta", "categoria", "seguidores", "seguidos", "enlace"])
                        for r in rows:
                            writer.writerow(r)
                    print(f"Archivo CSV guardado en {out_path}")
                    print(f"Tiempo total: {round(_t.time()-t0, 2)}s")
                    return
                except Exception as e:
                    print(f"No se pudo escribir CSV (.csv): {e}. Se imprimirá JSON.")
    elif args.command == "followers":
        import time as _t
        t0 = _t.time()
        scraper = BrowserInstagramScraper(config)
        data = scraper.get_followers_counts_for_followers(
            args.url,
            followers_limit=args.limit,
            page_size=args.page_size,
            chunk=args.chunk,
            delay_ms=args.delay_ms,
            retry_tries=args.retry_tries,
            retry_base_ms=args.retry_base_ms,
        )
        out_path = getattr(args, "output", None)
        followers_items = data.get("followers_of_followers", [])
        if out_path and out_path.suffix.lower() in {".xlsx", ".csv"}:
            rows = []
            print(f"Items scrapeados: {len(followers_items)}")
            for it in followers_items:
                username = it.get("username")
                followers = it.get("followers")
                print(f"{username}: {followers}")
                first_digit = None
                if isinstance(followers, int):
                    s = str(followers)
                    first_digit = int(s[0]) if s else None
                rows.append([username, followers if followers is not None else "", first_digit if first_digit is not None else ""])

            if out_path.suffix.lower() == ".xlsx":
                try:
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "followers"
                    ws.append(["username", "seguidores", "primer_digito"])
                    for r in rows:
                        ws.append(r)
                    out_path.parent.mkdir(parents=True, exist_ok=True)
                    wb.save(str(out_path))
                    print(f"Archivo Excel guardado en {out_path}")
                    print(f"Tiempo total: {round(_t.time()-t0, 2)}s")
                    return
                except Exception as e:
                    print(f"No se pudo escribir Excel (.xlsx): {e}. Se imprimirá JSON.")

            if out_path.suffix.lower() == ".csv":
                try:
                    import csv
                    out_path.parent.mkdir(parents=True, exist_ok=True)
                    with out_path.open("w", newline="", encoding="utf-8") as f:
                        writer = csv.writer(f)
                        writer.writerow(["username", "seguidores", "primer_digito"])
                        for r in rows:
                            writer.writerow(r)
                    print(f"Archivo CSV guardado en {out_path}")
                    print(f"Tiempo total: {round(_t.time()-t0, 2)}s")
                    return
                except Exception as e:
                    print(f"No se pudo escribir CSV (.csv): {e}. Se imprimirá JSON.")
    elif args.command == "legacy":
        scraper = InstagramScraper(config)
        if getattr(args, "login", False):
            scraper.login_if_available()
        data = scraper.get_profile_data(args.url, posts_limit=args.posts)
    else:
        parser.error("Comando no reconocido")
        return

    output = json.dumps(data, ensure_ascii=False, indent=2)
    print(output)

    out_path = getattr(args, "output", None)
    if out_path:
        if out_path.suffix.lower() == ".xlsx":
            pass
        else:
            out_path.parent.mkdir(parents=True, exist_ok=True)
            out_path.write_text(output, encoding="utf-8")
